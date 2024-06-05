import openpyxl
from datetime import datetime, timedelta

# Função para encontrar a próxima linha vazia em uma planilha
def encontrar_proxima_linha(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                return cell.row
    return sheet.max_row + 1


# Classe Livro
class Livro:
    def __init__(self, titulo, autor, genero, tipo, status):
        self.titulo = titulo
        self.autor = autor
        self.genero = genero
        self.tipo = tipo
        self.status = status

    def salvar(self, workbook):
        sheet = workbook['LIVROS']
        row = encontrar_proxima_linha(sheet)
        sheet.cell(row=row, column=1, value=self.titulo)
        sheet.cell(row=row, column=2, value=self.autor)
        sheet.cell(row=row, column=3, value=self.genero)
        sheet.cell(row=row, column=4, value=self.tipo)
        sheet.cell(row=row, column=5, value=self.status)
        workbook.save('POO.xlsx')

# Subclasse LivroFisico
class LivroFisico(Livro):
    def __init__(self, titulo, autor, genero, status='Disponível'):
        super().__init__(titulo, autor, genero, 'Físico', status)

# Subclasse LivroDigital
class LivroDigital(Livro):
    def __init__(self, titulo, autor, genero, status='Disponível'):
        super().__init__(titulo, autor, genero, 'Digital', status)

# Classe Usuario
class Usuario:
    def __init__(self, id, nome, cpf, senha, tipo):
        self.id = id
        self.nome = nome
        self.cpf = cpf
        self.senha = senha
        self.tipo = tipo

    def salvar(self, workbook):
        sheet = workbook['USUARIOS']
        row = encontrar_proxima_linha(sheet)
        sheet.cell(row=row, column=1, value=self.id)
        sheet.cell(row=row, column=2, value=self.nome)
        sheet.cell(row=row, column=3, value=self.cpf)
        sheet.cell(row=row, column=4, value=self.senha)
        sheet.cell(row=row, column=5, value=self.tipo)
        workbook.save('POO.xlsx')

# Classe Reserva
class Reserva:
    def __init__(self, titulo, tipo, id_usuario, data_reserva, data_devolucao=None):
        self.titulo = titulo
        self.tipo = tipo
        self.id_usuario = id_usuario
        self.data_reserva = data_reserva
        self.data_devolucao = data_devolucao

    def salvar(self, workbook):
        sheet = workbook['RESERVAS']
        row = encontrar_proxima_linha(sheet)
        sheet.cell(row=row, column=1, value=self.titulo)
        sheet.cell(row=row, column=2, value=self.tipo)
        sheet.cell(row=row, column=3, value=self.id_usuario)
        sheet.cell(row=row, column=4, value=self.data_reserva)
        sheet.cell(row=row, column=5, value=self.data_devolucao)
        workbook.save('POO.xlsx')

# Classe Sistema
class Sistema:
    def __init__(self, arquivo_excel):
        self.arquivo_excel = arquivo_excel
        self.workbook = openpyxl.load_workbook(arquivo_excel)

    def adicionar_livro(self, livro):
        livro.salvar(self.workbook)

    def adicionar_usuario(self, usuario):
        usuario.salvar(self.workbook)

    def adicionar_reserva(self, reserva):
        reserva.salvar(self.workbook)

    def listar(self, tipo):
        sheet = self.workbook[tipo.upper()]
        resultados = []
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=5, values_only=True):
            if any(cell is not None for cell in row):
                resultados.append(row)
        return resultados

    def excluir(self, tipo, identificador):
        sheet = self.workbook[tipo.upper()]
        col_index = 1  # Identificadores estão na coluna 1 (A)
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=5):
            if row[col_index-1].value == identificador:
                sheet.delete_rows(row[0].row)
                break
        self.workbook.save(self.arquivo_excel)

    def excluir_usuario(self, id_usuario):
        self.excluir('USUARIOS', id_usuario)

    def excluir_livro(self, titulo):
        self.excluir('LIVROS', titulo)

    def salvar(self):
        self.workbook.save(self.arquivo_excel)

    def emprestar_livro(self, titulo, id_usuario):
        sheet_livros = self.workbook['LIVROS']
        for row in sheet_livros.iter_rows(min_row=3, max_row=sheet_livros.max_row, min_col=1, max_col=5):
            if row[0].value == titulo and row[4].value == 'Disponível':
                row[4].value = 'Emprestado'
                self.workbook.save(self.arquivo_excel)
                data_reserva = datetime.now().date()
                data_devolucao = data_reserva + timedelta(days=15)
                self.adicionar_reserva(Reserva(titulo, row[3].value, id_usuario, data_reserva, data_devolucao))
                print(f'Livro {titulo} emprestado para o usuário {id_usuario}.')
                return
        print(f'Livro {titulo} não está disponível para empréstimo.')

    def devolver_livro(self, titulo, id_usuario):
        sheet_livros = self.workbook['LIVROS']
        sheet_reservas = self.workbook['RESERVAS']
        for row in sheet_livros.iter_rows(min_row=3, max_row=sheet_livros.max_row, min_col=1, max_col=5):
            if row[0].value == titulo and row[4].value == 'Emprestado':
                row[4].value = 'Disponível'
                self.workbook.save(self.arquivo_excel)
                for res_row in sheet_reservas.iter_rows(min_row=3, max_row=sheet_reservas.max_row, min_col=1, max_col=5):
                    if res_row[0].value == titulo and res_row[2].value == id_usuario and res_row[4].value is None:
                        res_row[4].value = datetime.now().date()
                        res_row[5].value = datetime.now().time().strftime('%H:%M:%S')
                        self.workbook.save(self.arquivo_excel)
                        print(f'Livro {titulo} devolvido pelo usuário {id_usuario}.')
                        return
        print(f'Livro {titulo} não foi encontrado como emprestado para o usuário {id_usuario}.')

    def pesquisa_avancada(self, **kwargs):
        sheet = self.workbook['LIVROS']
        resultados = []
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=5, values_only=True):
            match = True
            for key, value in kwargs.items():
                if key == 'titulo' and value.lower() not in row[0].lower():
                    match = False
                if key == 'autor' and value.lower() not in row[1].lower():
                    match = False
                if key == 'genero' and value.lower() not in row[2].lower():
                    match = False
                if key == 'tipo' and value.lower() != row[3].lower():
                    match = False
                if key == 'status' and value.lower() != row[4].lower():
                    match = False
            if match:
                resultados.append(row)
        return resultados

    def reservar_livro(self, titulo, id_usuario):
        sheet_livros = self.workbook['LIVROS']
        for row in sheet_livros.iter_rows(min_row=3, max_row=sheet_livros.max_row, min_col=1, max_col=5):
            if row[0].value == titulo and row[4].value == 'Disponível':
                row[4].value = 'Reservado'
                self.workbook.save(self.arquivo_excel)
                self.adicionar_reserva(Reserva(titulo, row[3].value, id_usuario, datetime.now().date()))
                print(f'Livro {titulo} reservado pelo usuário {id_usuario}.')
                return
        print(f'Livro {titulo} não está disponível para reserva.')

if __name__ == "__main__":
    sistema = Sistema('POO.xlsx')